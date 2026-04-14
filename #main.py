import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook  
# standard openpyxl code:

class ExcelDataProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Processor")
        self.root.geometry("400x400")
        self.file_path = None
        self.workbook = None

        #Open File Button
        self.open_button = tk.Button(root, text="Open Excel File", command=self.open_file)
        self.open_button.pack(pady=20)

        #Frame
        self.main_frame = tk.Frame(root)
        self.main_frame.pack(pady = 10 ,fill=tk.BOTH, expand=True)

        #Sheet Listbox
        self.sheet_frame = tk.Frame(self.main_frame)
        self.sheet_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.sheet_label = tk.Label(self.sheet_frame, text="Sheets")
        self.sheet_label.pack()
        self.sheet_listbox = tk.Listbox(self.sheet_frame, width=20)
        self.sheet_listbox.pack(fill=tk.BOTH, expand=True)
        self.sheet_listbox.bind("<<ListboxSelect>>", self.on_sheet_select)

        #Names list Section
        self.names_frame = tk.Frame(self.main_frame)
        self.names_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        self.names_label = tk.Label(self.names_frame, text="Names")
        self.names_label.pack()

        self.names_listbox = tk.Listbox(self.names_frame, width=40)
        self.names_listbox.pack(fill=tk.BOTH, expand=True)




    def open_file(self):
        file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            messagebox.showwarning("No File Selected", "Please select an Excel file to proceed.")
            return
        try: 
            self.workbook = load_workbook(filename=file_path, read_only=True)
            self.file_path = file_path
            
            self.sheet_listbox.delete(0, tk.END)  # Clear existing items in the listbox
            self.names_listbox.delete(0, tk.END)  # Clear existing items in the names listbox

            for sheet_name in self.workbook.sheetnames:
                self.sheet_listbox.insert(tk.END, sheet_name)

        except FileNotFoundError:
                messagebox.showerror("File Not Found", f"File '{file_path}' not found. Please check the file path and try again.")
        except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {e}")
             
    def on_sheet_select(self, event):
        if not self.workbook:
             return
        selection = self.sheet_listbox.curselection()
        if not selection:
             return
        
        selected_index = selection[0]
        sheet_name = self.sheet_listbox.get(selected_index)

        names = self.list_names(sheet_name)
        self.names_listbox.delete(0, tk.END)    


        if names: 
             for name in names:
                    self.names_listbox.insert(tk.END, name)
        else:
             self.names_listbox.insert(tk.END, "No names found in this sheet.")
   
   
    def list_names(self, sheet_name):
        try:
            if sheet_name not in self.workbook.sheetnames:
                messagebox.showerror("Sheet Not Found", f"Sheet '{sheet_name}' not found in the workbook.")
                return []

            sheet = self.workbook[sheet_name]
            names = []

            for row in sheet.iter_rows(min_row=4, values_only=True):
                name = row[0]
                if name is None:
                    continue
                if isinstance(name, str) and name.startswith("$"):
                    continue
                names.append(name)

            return names
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while processing the sheet: {e}")
            return []
         

#Run GUI

root = tk.Tk()
app = ExcelDataProcessor(root)
root.mainloop()
                 


    